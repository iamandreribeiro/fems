"""One-shot extractor (DEV-only; requires openpyxl).

Pulls the canonical baseline out of the v8 workbook and writes two committed,
runtime-friendly artefacts under ``src/fems/data`` — the runtime never reads the
.xlsx:

  * ``base_irradiacao_2025.json.gz`` — the 8760-hour climate year (read via stdlib).
  * ``catalog_seed.py`` — GENERATED typed literals for the catalog (equipamentos,
    geradores, tarifa). Reviewable, mypy-checked, no openpyxl at runtime.

Run:  uv run python scripts/_extract_from_xlsx.py
      (or: .venv/Scripts/python scripts/_extract_from_xlsx.py)
"""

from __future__ import annotations

import gzip
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "modelo_gestao_energia_fazenda_v8.xlsx"
DATA = ROOT / "src" / "fems" / "data"
CLIMA_OUT = DATA / "base_irradiacao_2025.json.gz"
SEED_OUT = DATA / "catalog_seed.py"

HOURS_IN_YEAR_2025 = 8760

AREA_ENUM = {
    "Escritório": "Area.ESCRITORIO",
    "Cozinha": "Area.COZINHA",
    "Quarto": "Area.QUARTO",
    "Irrigação/Lavoura": "Area.IRRIGACAO",
}
TIPO_GER_ENUM = {"Solar FV": "TipoGeracao.SOLAR_FV", "Eólica": "TipoGeracao.EOLICA"}
TIPO_HOR_ENUM = {"Ponta": "TipoHorario.PONTA", "Fora Ponta": "TipoHorario.FORA_PONTA"}

# --- Porte Grande (NÃO está na planilha v8) — valores PROPOSTOS, ajuste à vontade ---
# qtd_grande é derivada da Média por um fator; os geradores GRD são acréscimos.
FATOR_GRANDE = 2.0
GERADORES_GRANDE = [
    # (id, tipo_enum, pot_nominal_kwp, eficiencia_pct, ref_conversao, gen_max_kw, gen_min_kw)
    ("SOL-GRD", "TipoGeracao.SOLAR_FV", 100.0, 88.0, 1100.0, 100.0, 8.0),
    ("EOL-GRD", "TipoGeracao.EOLICA", 30.0, 88.0, 10.0, 30.0, 1.0),
]


def _qtd_grande(qtd_med: float) -> float:
    return float(math.floor(qtd_med * FATOR_GRANDE + 0.5))


def _grid(wb: Any, name: str) -> list[list[Any]]:
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def extract_clima(wb: Any) -> None:
    ghi: list[float] = []
    temp: list[float] = []
    vento: list[float] = []
    fp: list[float] = []
    for row in wb["Base_Irradiacao"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        ghi.append(float(row[3]))
        temp.append(float(row[4]))
        vento.append(float(row[5]))
        fp.append(float(row[6]))
    assert len(ghi) == HOURS_IN_YEAR_2025, f"esperado {HOURS_IN_YEAR_2025}, obtido {len(ghi)}"
    payload = {"ano": 2025, "n": len(ghi), "ghi": ghi, "temp": temp, "vento": vento, "fp": fp}
    CLIMA_OUT.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(CLIMA_OUT, "wt", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"))
    print(f"wrote {CLIMA_OUT} ({CLIMA_OUT.stat().st_size / 1024:.1f} KiB, {len(ghi)} h)")


def _fmt(x: float) -> str:
    return repr(float(x))


def generate_catalog_seed(wb: Any) -> None:
    equip_rows = [r for r in _grid(wb, "Config_Equipamentos")[1:] if r[0] is not None]
    ger_rows = [r for r in _grid(wb, "Config_Geracao")[1:] if r[0] is not None]
    tar_rows = [r for r in _grid(wb, "Tarifa")[1:] if r[0] is not None]

    lines: list[str] = []
    lines.append('"""GERADO por scripts/_extract_from_xlsx.py — NÃO editar à mão.')
    lines.append("")
    lines.append("Catálogo baseline extraído de modelo_gestao_energia_fazenda_v8.xlsx")
    lines.append("(Config_Equipamentos, Config_Geracao, Tarifa). Literais tipados; o")
    lines.append("runtime não lê o .xlsx.")
    lines.append('"""')
    lines.append("")
    lines.append("from fems.domain.configuration.enums import Area, TipoGeracao, TipoHorario")
    lines.append("from fems.domain.simulation.types import Equipamento, Gerador, TarifaHora")
    lines.append("")
    lines.append("EQUIPAMENTOS: list[Equipamento] = [")
    for r in equip_rows:
        perfil = tuple(float(x) for x in r[6:30])
        assert len(perfil) == 24, f"{r[0]}: perfil len {len(perfil)}"
        perfil_str = ", ".join(_fmt(x) for x in perfil)
        lines.append("    Equipamento(")
        lines.append(f"        id={r[0]!r},")
        lines.append(f"        area={AREA_ENUM[r[1]]},")
        lines.append(f"        equipamento={r[2]!r},")
        lines.append(f"        potencia_kw={_fmt(r[3])},")
        lines.append(f"        qtd_peq={_fmt(r[4])},")
        lines.append(f"        qtd_med={_fmt(r[5])},")
        lines.append(f"        qtd_grande={_fmt(_qtd_grande(float(r[5])))},")
        lines.append(f"        perfil=({perfil_str}),")
        lines.append("    ),")
    lines.append("]")
    lines.append("")
    lines.append("GERADORES: list[Gerador] = [")
    for r in ger_rows:
        lines.append("    Gerador(")
        lines.append(f"        id={r[0]!r},")
        lines.append(f"        tipo={TIPO_GER_ENUM[r[1]]},")
        lines.append(f"        pot_nominal_kwp={_fmt(r[2])},")
        lines.append(f"        eficiencia_pct={_fmt(r[3])},")
        lines.append(f"        ref_conversao={_fmt(r[4])},")
        lines.append(f"        gen_max_kw={_fmt(r[5])},")
        lines.append(f"        gen_min_kw={_fmt(r[6])},")
        lines.append("    ),")
    for gid, gtipo, kwp, efic, ref, gmax, gmin in GERADORES_GRANDE:
        lines.append("    Gerador(  # PROPOSTO (porte Grande, fora da planilha v8)")
        lines.append(f"        id={gid!r},")
        lines.append(f"        tipo={gtipo},")
        lines.append(f"        pot_nominal_kwp={_fmt(kwp)},")
        lines.append(f"        eficiencia_pct={_fmt(efic)},")
        lines.append(f"        ref_conversao={_fmt(ref)},")
        lines.append(f"        gen_max_kw={_fmt(gmax)},")
        lines.append(f"        gen_min_kw={_fmt(gmin)},")
        lines.append("    ),")
    lines.append("]")
    lines.append("")
    lines.append('TARIFA_AZUL_NOME = "AZUL_HOROSSAZONAL"')
    lines.append("TARIFA_AZUL: list[TarifaHora] = [")
    for r in tar_rows:
        hora = int(str(r[0]).split(":")[0])
        lines.append(
            f"    TarifaHora(hora={hora}, preco_kwh={_fmt(r[1])}, tipo={TIPO_HOR_ENUM[r[2]]}),"
        )
    lines.append("]")
    lines.append("")

    SEED_OUT.write_text("\n".join(lines), encoding="utf-8")
    print(
        f"wrote {SEED_OUT} ({len(equip_rows)} equip, {len(ger_rows)} ger, {len(tar_rows)} tarifa)"
    )


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    extract_clima(wb)
    generate_catalog_seed(wb)


if __name__ == "__main__":
    main()
